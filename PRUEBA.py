from openpyxl import load_workbook

# Cargar el archivo Excel
ruta_archivo = "C:\\Users\\Usuario\\Downloads\\DESCUENTOSCopia de Desc Avent Dic24.xlsx"
wb = load_workbook(ruta_archivo)
hoja = wb.active  # Selecciona la hoja activa o específica con wb['NombreHoja']

# Especificar la columna que deseas analizar (ejemplo: columna F)
columna = 'E'

# Crear una lista de los rangos combinados para evitar modificarla durante la iteración
rangos_combinados = list(hoja.merged_cells.ranges)

# Recorrer los rangos combinados en la hoja
for rango_combinado in rangos_combinados:  # Lista de celdas combinadas
    # Verificar si el rango pertenece a la columna especificada
    if rango_combinado.min_col == rango_combinado.max_col and rango_combinado.min_col == hoja[columna + '1'].column:
        # Obtener las coordenadas del rango combinado
        min_row, min_col, max_row, max_col = rango_combinado.bounds

        # Obtener el valor de la celda superior del rango combinado
        valor_celda = hoja.cell(row=min_row, column=min_col).value

        # Desagrupar la celda combinada
        hoja.unmerge_cells(rango_combinado.coord)

        # Copiar el valor hacia abajo en todas las celdas del rango
        for row in range(min_row, max_row + 1):
            hoja.cell(row=row, column=min_col).value = valor_celda

# Guardar el archivo modificado
wb.save('archivo_modificado.xlsx')
