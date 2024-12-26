import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook

class ExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Processor")
        self.root.geometry("900x500")
        self.root.configure(bg="#e0e0e0")
        
        # Estilos de ttk
        style = ttk.Style()
        style.theme_use("default")
        style.configure("TButton", font=("Arial", 12, "bold"), padding=10, foreground="white", background="#007bff")
        style.map("TButton", background=[("active", "#0056b3"), ("!disabled", "#007bff")])
        
        # Variables globales
        self.file_names = []
        self.file_checkboxes = []
        self.df_combined = pd.DataFrame()
        self.current_file_index = 0
        self.dfs = []
        self.header_line = None
        self.df = None
        self.selected_columns_order = []
        self.process_by_sheets = tk.BooleanVar()
        self.current_sheet_name = ""
        self.reference_df = None  # DataFrame de referencia
        
        # Elementos de la GUI
        tk.Label(self.root, text="Escoge una o más opciones", font=("Arial", 16), bg="#e0e0e0").pack(pady=10)
        self.process_by_sheets_checkbox = ttk.Checkbutton(self.root, text="Procesar por hojas", variable=self.process_by_sheets)
        self.process_by_sheets_checkbox.pack(pady=5)
        self.upload_button = ttk.Button(self.root, text="Subir archivos", command=self.upload_and_list_files)
        self.upload_button.pack(pady=20)

    def upload_and_list_files(self):
        self.file_names = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not self.file_names:
            return
        if self.process_by_sheets.get():
            for file in self.file_names:
                xls = pd.ExcelFile(file)
                sheet_checkboxes = [tk.BooleanVar() for _ in xls.sheet_names]
                sheet_selection_box = tk.Toplevel(self.root)
                sheet_selection_box.configure(bg="#e0e0e0")
                for var, sheet in zip(sheet_checkboxes, xls.sheet_names):
                    ttk.Checkbutton(sheet_selection_box, text=sheet, variable=var).pack(anchor='w', padx=10, pady=5)
                button_process_sheets = ttk.Button(sheet_selection_box, text=f"Procesar hojas de {file}", 
                                                  command=lambda f=file, s_cb=sheet_checkboxes, w=sheet_selection_box: self.process_sheets(f, s_cb, w))
                button_process_sheets.pack(pady=10)
        else:
            file_selection_box = tk.Toplevel(self.root)
            file_selection_box.configure(bg="#e0e0e0")
            self.file_checkboxes = [tk.BooleanVar() for _ in self.file_names]
            for var, name in zip(self.file_checkboxes, self.file_names):
                ttk.Checkbutton(file_selection_box, text=name, variable=var).pack(anchor='w', padx=10, pady=5)
            button_process_files = ttk.Button(file_selection_box, text="Procesar archivos seleccionados", command=lambda: self.process_files(file_selection_box))
            button_process_files.pack(pady=10)
            

    def process_sheets(self, file, sheet_checkboxes, window):
        selected_sheets = [sheet for var, sheet in zip(sheet_checkboxes, pd.ExcelFile(file).sheet_names) if var.get()]
        if not selected_sheets:
            selected_sheets = pd.ExcelFile(file).sheet_names
        self.current_file_index = 0
        self.dfs = [(file, sheet) for sheet in selected_sheets]
        window.destroy()
        self.process_next_file()

    def process_files(self, window):
        selected_files = [name for var, name in zip(self.file_checkboxes, self.file_names) if var.get()]
        if not selected_files:
            selected_files = self.file_names
        self.current_file_index = 0
        self.dfs = [(file, None) for file in selected_files]
        window.destroy()
        self.process_next_file()

    def process_next_file(self):
        if self.current_file_index < len(self.dfs):
            current_file, current_sheet = self.dfs[self.current_file_index]
            self.current_sheet_name = current_sheet
            self.header_line = tk.IntVar(value=1)
            header_line_label = tk.Label(self.root, text=f"Línea de encabezado para {current_file} - Hoja: {current_sheet}:", bg="#e0e0e0")
            header_line_label.pack()
            header_line_entry = tk.Entry(self.root, textvariable=self.header_line)
            header_line_entry.pack()
            button_read_header = ttk.Button(self.root, text=f"Leer encabezados de {current_file} - Hoja: {current_sheet}", 
                                           command=lambda: self.read_header(current_file, current_sheet))
            button_read_header.pack(pady=10)
        else:
            button_generate_combined = ttk.Button(self.root, text="Generar archivo combinado", command=self.generate_combined_file)
            button_generate_combined.pack(pady=20)
            button_add_reference = ttk.Button(self.root, text="Agregar columna desde archivo de referencia", command=self.upload_reference_file)
            button_add_reference.pack(pady=10)

    def read_header(self, file_name, sheet_name):
        header_row = self.header_line.get() - 1
        if sheet_name is None:
            self.df = pd.read_excel(file_name, engine='openpyxl', header=header_row, dtype=str)
        else:
            self.df = pd.read_excel(file_name, sheet_name=sheet_name, engine='openpyxl', header=header_row, dtype=str)
        self.df.reset_index(drop=True, inplace=True)

        columns_window = tk.Toplevel(self.root)
        columns_window.geometry("900x400")
        columns_window.configure(bg="#e0e0e0")
        columns_window.title(f"Cabeceras del archivo {file_name} - Hoja: {sheet_name}")

        frame_canvas = tk.Frame(columns_window)
        frame_canvas.pack(fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(frame_canvas, bg="#e0e0e0")
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar_y = tk.Scrollbar(frame_canvas, orient="vertical", command=canvas.yview)
        scrollbar_y.pack(side=tk.RIGHT, fill="y")

        scrollbar_x = tk.Scrollbar(frame_canvas, orient="horizontal", command=canvas.xview)
        scrollbar_x.pack(side=tk.BOTTOM, fill="x")

        canvas.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

        frame_scrollable = tk.Frame(canvas, bg="#e0e0e0")
        canvas.create_window((0, 0), window=frame_scrollable, anchor="nw")

        def on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        frame_scrollable.bind("<Configure>", on_frame_configure)

        columns_label = tk.Label(frame_scrollable, text=f"Cabeceras del archivo {file_name} - Hoja: {sheet_name}:", bg="#e0e0e0")
        columns_label.grid(row=0, column=0, columnspan=2)

        self.checkboxes = []
        self.order_inputs = []

        for row, col in enumerate(self.df.columns, start=1):
            var = tk.BooleanVar()
            cb = ttk.Checkbutton(frame_scrollable, text=col, variable=var)
            cb.grid(row=row, column=0, sticky="w", padx=3, pady=3)
            cb.grid(row=row, column=1, sticky="w", padx=3, pady=3) # Reducir espaciado horizontal
            cb.var = var
            self.checkboxes.append(cb)

            order_var = tk.IntVar(value=0)
            order_entry = tk.Entry(frame_scrollable, textvariable=order_var, width=5)  # Ajustar el tamaño del input
            order_entry.grid(row=row, column=1, padx=5, pady=3)  # Reducir espaciado vertical
            self.order_inputs.append(order_var)

        button_generate = ttk.Button(frame_scrollable, text="Agregar columnas seleccionadas", command=lambda: self.select_columns(columns_window))
        button_generate.grid(row=row+1, column=0, columnspan=2, pady=10)


    def select_columns(self, window):
        self.selected_columns_order = sorted([(order.get(), index) for index, order in enumerate(self.order_inputs) if order.get() > 0])
        window.destroy()
        selected_columns = [self.df.iloc[:, idx[1]].copy() for idx in self.selected_columns_order]
        new_df = pd.concat(selected_columns, axis=1)
        new_df.columns = [f"Column_{i+1}" for i in range(len(new_df.columns))]
        if self.df_combined.empty:
            self.df_combined = new_df
        else:
            self.df_combined = pd.concat([self.df_combined, new_df], ignore_index=True)
        messagebox.showinfo("Info", "Columnas agregadas correctamente.")
        self.current_file_index += 1
        self.process_next_file()















    def upload_reference_file(self):
        reference_file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if reference_file:
            # Cargar el archivo con openpyxl
            wb = load_workbook(reference_file)
            sheet = wb.active  # Puedes permitir al usuario seleccionar la hoja si es necesario

            # Descombinar celdas en la hoja activa
            merged_ranges = list(sheet.merged_cells.ranges)  # Copiamos la lista porque se modificará
            for merged_range in merged_ranges:
                # Obtener el rango combinado
                min_row, min_col, max_row, max_col = merged_range.bounds
                merged_value = sheet.cell(row=min_row, column=min_col).value  # Obtener el valor de la celda combinada


                # Descombinar las celdas
                sheet.unmerge_cells(merged_range.coord)

                # Rellenar las celdas con el valor combinado
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        cell = sheet.cell(row=row, column=col)  # Acceder a la celda real
                        cell.value = merged_value  # Asignar el valor

            # Guardar los cambios en un archivo temporal para leerlo con pandas
            temp_file = "temp_reference_file.xlsx"
            wb.save(temp_file)

            # Leer el archivo procesado con pandas
            self.reference_df = pd.read_excel(temp_file, dtype=str)











            self.reference_window = tk.Toplevel(self.root)
            self.reference_window.title("Columnas de referencia")
            tk.Label(self.reference_window, text="Selecciona la columna de EAN y la columna a agregar:", bg="#e0e0e0").pack(pady=10)
            self.id_column_var = tk.StringVar()
            self.column_to_add_var = tk.StringVar()
            id_label = tk.Label(self.reference_window, text="Columna de EAN:", bg="#e0e0e0")
            id_label.pack(pady=5)
            id_entry = tk.Entry(self.reference_window, textvariable=self.id_column_var)
            id_entry.pack(pady=5)

            column_label = tk.Label(self.reference_window, text="Columna a agregar:", bg="#e0e0e0")
            column_label.pack(pady=5)
            column_entry = tk.Entry(self.reference_window, textvariable=self.column_to_add_var)
            column_entry.pack(pady=5)

            button_add_column = ttk.Button(self.reference_window, text="Agregar columna", command=self.add_reference_column)
            button_add_column.pack(pady=10)
        return reference_file

    def add_reference_column(self):
        try:
            # Obtén los índices de las columnas especificadas por el usuario
            id_column_index = int(self.id_column_var.get().strip())-1  # Índice de la columna de ID en el archivo de referencia
            column_to_add_index = int(self.column_to_add_var.get().strip())-1  # Índice de la columna a agregar desde el archivo de referencia

            # Verifica que los índices sean válidos
            if id_column_index < 0 or id_column_index >= self.reference_df.shape[1]:
                messagebox.showerror("Error", f"El índice de ID '{id_column_index}' está fuera de rango.")
                return

            if column_to_add_index < 0 or column_to_add_index >= self.reference_df.shape[1]:
                messagebox.showerror("Error", f"El índice de la columna '{column_to_add_index}' está fuera de rango.")
                return

            # Obtén los nombres de las columnas basados en los índices
            id_column = self.reference_df.columns[id_column_index]
            column_to_add = self.reference_df.columns[column_to_add_index]

            # Realiza el merge basándote en las columnas especificadas
            self.df_combined = self.df_combined.merge(
                self.reference_df[[id_column, column_to_add]],  # Solo incluye la columna de ID y la columna a agregar
                left_on=self.df_combined.columns[0],  # Primera columna del archivo combinado (por ejemplo, "Column_1")
                right_on=id_column,  # Columna de ID del archivo de referencia
                how='left'
            )

            # Elimina la columna de ID del archivo de referencia si no se desea conservar
            self.df_combined.drop(columns=[id_column], inplace=True)

            # Elimina las filas donde la columna principal del archivo combinado tiene valores nulos
            self.df_combined.dropna(subset=[self.df_combined.columns[0]], inplace=True)

            messagebox.showinfo("Éxito", f"La columna '{column_to_add}' se ha agregado correctamente.")
        except ValueError:
            messagebox.showerror("Error", "Por favor, ingresa índices válidos para las columnas.")
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error durante la combinación: {str(e)}")

    def generate_combined_file(self):
        if not self.df_combined.empty:
            # Crear una ventana para renombrar las columnas
            rename_window = tk.Toplevel(self.root)
            rename_window.title("Renombrar columnas")
            rename_window.configure(bg="#e0e0e0")
            
            tk.Label(rename_window, text="Renombra las columnas antes de guardar:", bg="#e0e0e0").pack(pady=10)

            # Crear entradas para renombrar columnas
            new_column_names = []
            for idx, col in enumerate(self.df_combined.columns):
                frame = tk.Frame(rename_window, bg="#e0e0e0")
                frame.pack(pady=5, padx=10, fill=tk.X)

                tk.Label(frame, text=f"Columna {idx + 1}: {col}", bg="#e0e0e0").pack(side=tk.LEFT)
                new_name_var = tk.StringVar(value=col)
                tk.Entry(frame, textvariable=new_name_var, width=30).pack(side=tk.RIGHT, padx=5)
                new_column_names.append(new_name_var)

            def save_file():
                # Renombrar las columnas
                renamed_columns = [var.get() for var in new_column_names]
                self.df_combined.columns = renamed_columns

                # Guardar el archivo
                output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("XLSX files", "*.xlsx")])
                if output_file:
                    self.df_combined.to_excel(output_file, index=False, header=True)
                    messagebox.showinfo("Éxito", f"Archivo combinado guardado en: {output_file}")
                    rename_window.destroy()

            # Botón para guardar
            ttk.Button(rename_window, text="Guardar archivo", command=save_file).pack(pady=10)
        else:
            messagebox.showerror("Error", "No hay datos para combinar.")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelProcessorApp(root)
    root.mainloop()